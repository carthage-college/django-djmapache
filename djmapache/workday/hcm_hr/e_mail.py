import os
import sys
# import pysftp
import csv
import argparse
import logging
from logging.handlers import SMTPHandler

# prime django
import django

# django settings for shell environment
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "djmapache.settings.shell")
# django.setup()

# django settings for script
import openpyxl
from openpyxl.utils.cell import get_column_letter

from openpyxl import Workbook
from django.conf import settings
from djimix.core.utils import get_connection, xsql
from djmapache.workday.hcm_hr.utilities import fn_write_email_cl_header, \
    fn_write_clean_file, fn_get_id


# informix environment
os.environ['INFORMIXSERVER'] = settings.INFORMIXSERVER
os.environ['DBSERVERNAME'] = settings.DBSERVERNAME
os.environ['INFORMIXDIR'] = settings.INFORMIXDIR
os.environ['ODBCINI'] = settings.ODBCINI
os.environ['ONCONFIG'] = settings.ONCONFIG
os.environ['INFORMIXSQLHOSTS'] = settings.INFORMIXSQLHOSTS
os.environ['LD_LIBRARY_PATH'] = settings.LD_LIBRARY_PATH
os.environ['LD_RUN_PATH'] = settings.LD_RUN_PATH

# normally set as 'debug" in SETTINGS
DEBUG = settings.INFORMIX_DEBUG

# set up command-line options
desc = """
    Upload ADP data to CX
"""
parser = argparse.ArgumentParser(description=desc)

parser.add_argument(
    "--test",
    action='store_true',
    help="Dry run?",
    dest="test"
)
parser.add_argument(
    "-d", "--database",
    help="database name.",
    dest="database"
)


def fn_clear_sheet(csv_output, new_xl_file, sheet):
    wb_obj = openpyxl.load_workbook(csv_output + new_xl_file)
    this_sheet = wb_obj[sheet]
    print(this_sheet)
    row_count = this_sheet.max_row
    col_count = this_sheet.max_column
    # print(row_count)
    # print(col_count)
    col = get_column_letter(col_count)
    print(col)

    for row in this_sheet['A3:'+ col + str(row_count)]:
        # print(row)
        for cell in row:
            # print(cell.value)
            cell.value = ""
            # print(cell.value)

    wb_obj.save(csv_output + new_xl_file)


def fn_insert_xl(ct, workerid, email_typ, email_addr, public,
                 email_csv_output, new_xl_file):

    try:
        wb_obj = openpyxl.load_workbook(email_csv_output + new_xl_file)
        # # print(wb_obj.sheetnames)
        this_sheet = wb_obj['Email']
        # print(this_sheet)

        this_sheet.cell(row=ct, column=1).value = workerid
        this_sheet.cell(row=ct, column=2).value = email_typ
        this_sheet.cell(row=ct, column=3).value = email_addr
        this_sheet.cell(row=ct, column=4).value = public

        wb_obj.save(email_csv_output + new_xl_file)

    except Exception as e:
        print("Error in email_rec.py fn_ins_xl , Error = " + repr(e))
    # fn_write_error("Error in email_rec.py - fn_get_id: "
    #                + repr(e))
    # fn_send_mail(settings.ADP_TO_EMAIL, settings.ADP_FROM_EMAIL,
    #          "Error in email_rec.py fn_get_id, Error = " + repr(e),
    #          "Error in email_rec.py fn_get_id")

def main():

    ##########################################################################
    # ==> python email_rec.py --database=train --test
    # ==> python email_rec.py --database=cars
    ##########################################################################

    # # Defines file names and directory location
    # if test:
    email_csv_output = "/home/dsullivan/djmapache/djmapache/workday/hcm_hr/" \
                            "clean_data/"
    email_csv_input = "/home/dsullivan/djmapache/djmapache/workday/hcm_hr/" \
                            "raw_data/"
    # else:
    # email_csv_output = settings.email_csv_output
    # print(email_csv_output)

    raw_email_file = email_csv_input + "email.csv"
    new_email_file = email_csv_output + "email_cln.csv"
    new_xl_file = "Worker Data.xlsx"

    try:
        # set global variable
        global EARL
        # determines which database is being called from the command line
        if database == 'cars':
            EARL = settings.INFORMIX_ODBC
        if database == 'train':
            EARL = settings.INFORMIX_ODBC_TRAIN
        else:
            # # this will raise an error when we call get_engine()
            # below but the argument parser should have taken
            # care of this scenario and we will never arrive here.
            EARL = None
            # establish database connection
        # print(EARL)

        # Pull the file from the ADP FTP site
        # execute sftp code in production only
        # if not test:
        #     file_download()

        # Create the new csv file for the formatted data
        fn_write_email_cl_header(new_email_file)

        # Read the raw file and work on the formatting
        # print(raw_email_file)
        # "Associate ID", "File Number", "Carthage ID #",
        # "Personal Contact: Personal Email",
        # "Personal Contact: Use Personal Email for Notification", \
        # "Work Contact: Use Work Email for Notification",
        # "Work Contact: Work Email"

        fn_clear_sheet(email_csv_output, new_xl_file, 'Email')

        with open(raw_email_file, 'r') as f:
            d_reader = csv.DictReader(f, delimiter=',')
            ct = 2  # Leave at 1 to skip the header row...
            r = 0
            # Write to Excel spreadsheet

            for row in d_reader:
                # print("+++++++++++++++++++++")
                # print(row)
                r = r+1
                # print("ROW " + str(r))
                # print("Inserts " + str(ct))

                """Student workers and a few others typically don't have the
                    Carthage ID Custom field populated  May have to force it
                    here"""
                if row['Carthage ID #'] == "":
                    print('No Carth ID  ' + row['File Number'])
                    if row['File Number'] == "":
                        print("No id or file number")
                        pass
                    else:
                        # print(str(row['File Number']))
                        workerid = fn_get_id(str(row['File Number']), EARL,
                                settings.INFORMIX_DEBUG)
                        print(workerid)
                else:
                    workerid = row['Carthage ID #']

                if len(row['Personal Contact: Personal Email']) != 0:
                    email_addr = row['Personal Contact: Personal Email']
                    fn_write_clean_file(new_email_file, [workerid, "Home",
                         email_addr,  "No"])
                    ct = ct + 1
                    fn_insert_xl(ct, workerid, "Home", email_addr, "No",
                                 email_csv_output, new_xl_file)

                if len(row['Work Contact: Work Email']) != 0:
                    email_addr = row['Work Contact: Work Email']
                    fn_write_clean_file(new_email_file, [workerid, "Work",
                                                       email_addr, "Yes"])
                    ct = ct + 1
                    fn_insert_xl(ct, workerid, "Work", email_addr, "Yes",
                                 email_csv_output, new_xl_file)

    except Exception as e:
        print("Error in email_rec.py, Error = " + repr(e))
        # fn_write_error("Error in email_rec.py - Main: "
        #                + repr(e))
        # fn_send_mail(settings.ADP_TO_EMAIL, settings.ADP_FROM_EMAIL,
        #          "Error in email_rec.py, Error = " + repr(e),
        #          "Error in email_rec.py")

if __name__ == "__main__":
    args = parser.parse_args()
    test = args.test
    database = args.database

    if not database:
        print("mandatory option missing: database name\n")
        exit(-1)
    else:
        database = database.lower()

    if database != 'cars' and database != 'train' and database != 'sandbox':
        print("database must be: 'cars' or 'train' or 'sandbox'\n")
        parser.print_help()

    sys.exit(main())
