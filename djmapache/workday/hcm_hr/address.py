import os
import sys
# import pysftp
import csv
import argparse
import logging
from logging.handlers import SMTPHandler
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "djmapache.settings.shell")
django.setup()
from django.conf import settings


import openpyxl
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

from django.conf import settings
from djimix.core.utils import get_connection, xsql
from djmapache.workday.hcm_hr.utilities import fn_format_country, \
    fn_write_addr_cl_header, fn_write_clean_file, fn_get_id, \
    fn_format_cx_country


# informix environment
os.environ['INFORMIXSERVER'] = settings.INFORMIXSERVER
os.environ['DBSERVERNAME'] = settings.DBSERVERNAME
os.environ['INFORMIXDIR'] = settings.INFORMIXDIR
os.environ['ODBCINI'] = settings.ODBCINI
os.environ['ONCONFIG'] = settings.ONCONFIG
os.environ['INFORMIXSQLHOSTS'] = settings.INFORMIXSQLHOSTS
os.environ['LD_LIBRARY_PATH'] = settings.LD_LIBRARY_PATH
os.environ['LD_RUN_PATH'] = settings.LD_RUN_PATH

# # normally set as 'debug" in SETTINGS
# DEBUG = settings.INFORMIX_DEBUG

# set up command-line options
desc = """
    Upload ADP data to CX
"""
parser = argparse.ArgumentParser(description=desc)

parser.add_argument(
    "--test",
    action='store_true',
    help="Dry run?",
    dest="test"
)
parser.add_argument(
    "-d", "--database",
    help="database name.",
    dest="database"
)


def fn_clear_sheet(csv_output, new_xl_file, sheet):
    wb_obj = openpyxl.load_workbook(csv_output + new_xl_file)
    this_sheet = wb_obj[sheet]
    print(this_sheet)
    row_count = this_sheet.max_row
    col_count = this_sheet.max_column
    # print(row_count)
    # print(col_count)
    col = get_column_letter(col_count)
    print(col)

    for row in this_sheet['A3:'+ col + str(row_count)]:
        # print(row)
        for cell in row:
            # print(cell.value)
            cell.value = ""
            # print(cell.value)

    wb_obj.save(csv_output + new_xl_file)


def fn_insert_xl(ct, workerid, cntry, addr_typ, addr_use, region, subregion,
                 city, city_subdv, postal_cod, addr1, addr2, addr3, addr4,
                 rem_ee,
                 csv_output, new_xl_file):
    try:
        wb_obj = openpyxl.load_workbook(csv_output
                                        + new_xl_file)
        # this may be the syntax to open a pwd protected xls..
        # wb_obj.protection.password = ''
        # After a fair amount of research, it seems that there is no easy
        # way to use a password withing Python to pass to an Excel file
        # May be easiest just to remove the password protection and re-enter
        # it when the work is done.

        # # print(wb_obj.sheetnames)
        this_sheet = wb_obj['Address']
        # print(this_sheet)
        this_sheet.cell(row=ct, column=1).value = workerid
        this_sheet.cell(row=ct, column=2).value = cntry
        this_sheet.cell(row=ct, column=3).value = addr_typ
        this_sheet.cell(row=ct, column=4).value = addr_use
        this_sheet.cell(row=ct, column=5).value = region
        this_sheet.cell(row=ct, column=6).value = subregion
        this_sheet.cell(row=ct, column=7).value = city

        this_sheet.cell(row=ct, column=8).value = city_subdv
        this_sheet.cell(row=ct, column=9).value = postal_cod
        this_sheet.cell(row=ct, column=10).value = addr1
        this_sheet.cell(row=ct, column=11).value = addr2
        this_sheet.cell(row=ct, column=12).value = addr3
        this_sheet.cell(row=ct, column=13).value = addr4
        this_sheet.cell(row=ct, column=14).value = rem_ee

        wb_obj.save(csv_output + new_xl_file)

    except Exception as e:
        print("Error in address.py fn_ins_xl , Error = " + repr(e))
    # fn_write_error("Error in address.py - fn_get_id: "
    #                + repr(e))
    # fn_send_mail(settings.ADP_TO_EMAIL, settings.ADP_FROM_EMAIL,
    #          "Error in address.py fn_get_id, Error = " + repr(e),
    #          "Error in address.py fn_get_id")

def main():

    ##########################################################################
    # ==> python phone_rec.py --database=train --test
    # ==> python phone_rec.py --database=cars
    ##########################################################################

    # # Defines file names and directory location
    # if test:
    csv_output = "/home/dsullivan/djmapache/djmapache/workday/hcm_hr/" \
                            "clean_data/"
    csv_input = "/home/dsullivan/djmapache/djmapache/workday/hcm_hr/" \
                             "raw_data/"
    # else:
    # csv_output = settings.csv_output
    # print(csv_output)

    # For testing use last file
    # new_file = csv_output + "ADPtoCXLast.csv"
    raw_file = csv_input + "Addresses.csv"
    new_file = csv_output + "address_cln.csv"
    new_xl_file = "Worker Data.xlsx"

    fn_clear_sheet(csv_output, new_xl_file, 'Address')

    try:
        # set global variable
        global EARL
        # determines which database is being called from the command line
        if database == 'cars':
            EARL = settings.INFORMIX_ODBC
        if database == 'train':
            EARL = settings.INFORMIX_ODBC_TRAIN
        else:
            # # this will raise an error when we call get_engine()
            # below but the argument parser should have taken
            # care of this scenario and we will never arrive here.
            EARL = None
            # establish database connection
        # print(EARL)

        # Pull the file from the ADP FTP site
        # execute sftp code in production only
        # if not test:
        #     file_download()

        # Create the new csv file for the formatted data
        fn_write_addr_cl_header(new_file)

        # Read the raw file and work on the formatting
        # print(raw_file)
        with open(raw_file, 'r') as f:
            d_reader = csv.DictReader(f, delimiter=',')
            ct = 2  # Leave at 1 to skip the header row...
            r = 0
            # Write to Excel spreadsheet

            for row in d_reader:
                # print("+++++++++++++++++++++")
                # print(row)
                r = r+1
                # print("ROW " + str(r))
                # print("Inserts " + str(ct))

                """Student workers and a few others typically 
                don't have the
                    Carthage ID Custom field populated  May have 
                    to force it
                    here"""
                if row['Carthage ID #'] == "":
                #     print('No Carth ID  ' + row['File Number'])
                    if row['File Number'] == "":
                        print("No id or file number")
                        pass
                    else:
                        # print(str(row['File Number']))
                        workerid = fn_get_id(str(row['File Number']), EARL,
                                            settings.INFORMIX_DEBUG)
                        # print(workerid)
                else:
                    workerid = row['Carthage ID #']
                    # print(workerid)

                """Format country and phone - I don't know if our 
                ADP  data
                    has any phone numbers that aren't in the US 
                    format"""

                cntry = fn_format_country(row['Primary Address: Country Code'])
                # cntry = row['Primary Address: Country Code']
                addr_typ = "Primary Home"
                addr_use = "Other - Home"
                region = row["Primary Address: State / Territory Code"]
                subregion = ""
                city = row["Primary Address: City"]
                city_subdv = ""
                postal_cod = row["Primary Address: Zip / Postal Code"]
                addr1 = row["Primary Address: Address Line 1"]
                addr2 = row["Primary Address: Address Line 2"]
                addr3 = ""
                addr4 = ""

                # Not sure how to populate this...
                remote_EE = ""

                # print(workerid, cntry, addr_typ, addr_use, region, city,
                #       postal_cod, addr1, addr2)

                if row["Primary Address: Country Code"] == 'USA':
                    fn_write_clean_file(new_file, [workerid + ',' + cntry + ','
                        + addr_typ + ',' + addr_use + ',' + region + ','
                        + subregion + ',' + city + ',' + city_subdv + ','
                        + postal_cod + ',' + addr1 + ',' + addr2 + ','
                        + addr3 + ',' + addr4 + ',' + remote_EE])
                    ct = ct + 1

                    fn_insert_xl(ct, workerid, cntry, addr_typ, addr_use,
                        region, subregion, city, city_subdv, postal_cod,
                        addr1, addr2, addr3, addr4,  remote_EE,
                        csv_output, new_xl_file)


        """Because we can't use the ADP info for student workers, we need to get 
            that info from cx"""
        # Get the student data via SQL query
        ct = ct + 1

        stuquery = '''select CR.adp_associate_id, CR.adp_id, JR.id, 
                trim(IR.ctry), trim(IR.st), trim(IR.city), trim(IR.addr_line1), 
                trim(IR.addr_line2), trim(IR.addr_line3), trim(IR.zip), 
                JR.hrpay, JR.beg_date, JR.end_date, JR.tpos_no, JR.job_title, 
                JR.supervisor_no, JR.title_rank  
                from job_rec JR
                LEFT JOIN id_rec IR
                ON JR.id = IR.id
                join cvid_rec CR
                on CR.cx_id = JR.id
                where (JR.end_date is null  or JR.end_date > TODAY)
                and JR.hrpay = 'DPW'
                limit 10
            '''

        print(stuquery)

        connection = get_connection(EARL)
        with connection:
            data_result = xsql(stuquery, connection).fetchall()
            ret = list(data_result)
            for i in ret:
                workerid = str(i[2])
                cntry = fn_format_cx_country(i[3])
                # cntry = row['Primary Address: Country Code']
                addr_typ = "Primary Home"
                addr_use = "Other - Home"
                region = i[4]
                subregion = ""
                city = i[5]
                city_subdv = ""
                postal_cod = i[9]
                addr1 = i[6]
                addr2 = i[7]
                addr3 = i[8]
                addr4 = ""
                remote_EE = ""

                """Foreign Addresses will be a challenge...
                   How to validate all the options?
                Argentina: requires country, City, Addr1, Postal code
                Cameroon: requires country, City, Addr1 -NO POSTAL CODE
                China: requires country, Region, Addr1, Postal code
                Senegal: requires country, NO REGION, City, Addr1, Postal code
                """

                print(ct)
                print(workerid + ',' + cntry + ','
                     + addr_typ + ',' + addr_use + ',' + region + ','
                     + subregion + ',' + city + ',' + city_subdv + ','
                     + postal_cod + ',' + addr1 + ',' + addr2 + ','
                     + addr3 + ',' + addr4 + ',' + remote_EE)

                fn_write_clean_file(new_file, [workerid + ',' + cntry + ','
                        + addr_typ + ',' + addr_use + ',' + region + ','
                        + subregion + ',' + city + ',' + city_subdv + ','
                        + postal_cod + ',' + addr1 + ',' + addr2 + ','
                        + addr3 + ',' + addr4 + ',' + remote_EE])

                fn_insert_xl(ct, workerid, cntry, addr_typ, addr_use,
                    region, subregion, city, city_subdv, postal_cod,
                    addr1, addr2, addr3, addr4,  remote_EE,
                    csv_output, new_xl_file)
                ct = ct + 1

        # Loop through and append data to csv and/or workbook

    except Exception as e:
        print("Error in address.py, Error = " + repr(e))
        # fn_write_error("Error in address.py - Main: "
        #                + repr(e))
        # fn_send_mail(settings.ADP_TO_EMAIL, settings.ADP_FROM_EMAIL,
        #          "Error in address.py, Error = " + repr(e),
        #          "Error in address.py")

if __name__ == "__main__":
    args = parser.parse_args()
    test = args.test
    database = args.database

    if not database:
        print("mandatory option missing: database name\n")
        exit(-1)
    else:
        database = database.lower()

    if database != 'cars' and database != 'train' and database != 'sandbox':
        print("database must be: 'cars' or 'train' or 'sandbox'\n")
        parser.print_help()

    sys.exit(main())
