import os
import sys
# import pysftp
import csv
import argparse
import logging
from logging.handlers import SMTPHandler
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "djmapache.settings.shell")
django.setup()
from django.conf import settings

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from django.conf import settings
from djimix.core.utils import get_connection, xsql
from djmapache.workday.hcm_hr.utilities import fn_format_country, \
    fn_format_phone, fn_write_phone_cl_header, fn_write_clean_file, \
    fn_get_id, fn_format_cx_country, fn_format_cx_phone

# informix environment
os.environ['INFORMIXSERVER'] = settings.INFORMIXSERVER
os.environ['DBSERVERNAME'] = settings.DBSERVERNAME
os.environ['INFORMIXDIR'] = settings.INFORMIXDIR
os.environ['ODBCINI'] = settings.ODBCINI
os.environ['ONCONFIG'] = settings.ONCONFIG
os.environ['INFORMIXSQLHOSTS'] = settings.INFORMIXSQLHOSTS
os.environ['LD_LIBRARY_PATH'] = settings.LD_LIBRARY_PATH
os.environ['LD_RUN_PATH'] = settings.LD_RUN_PATH

# # normally set as 'debug" in SETTINGS
# DEBUG = settings.INFORMIX_DEBUG

# set up command-line options
desc = """
    Upload ADP data to CX
"""
parser = argparse.ArgumentParser(description=desc)

parser.add_argument(
    "--test",
    action='store_true',
    help="Dry run?",
    dest="test"
)
parser.add_argument(
    "-d", "--database",
    help="database name.",
    dest="database"
)

"""NOTE:  Doesn't make sense to try to modularize the excel functions because
    each worksheet has so many specific columns with formatting  Will need 
    openpyxl in each script"""

def fn_clear_sheet(csv_output, new_xl_file, sheet):
    wb_obj = openpyxl.load_workbook(csv_output + new_xl_file)
    this_sheet = wb_obj[sheet]
    print(this_sheet)
    row_count = this_sheet.max_row
    col_count = this_sheet.max_column
    # print(row_count)
    # print(col_count)
    col = get_column_letter(col_count)
    print(col)

    for row in this_sheet['A3:'+ col + str(row_count)]:
        # print(row)
        for cell in row:
            # print(cell.value)
            cell.value = ""
            # print(cell.value)

    wb_obj.save(csv_output + new_xl_file)

def fn_insert_xl(ct, workerid, phon_typ, cntry, intlcode, area, phon, public,
                 csv_output, new_xl_file):

    try:
        wb_obj = openpyxl.load_workbook(csv_output
                                        + new_xl_file)
        # this may be the syntax to open a pwd protected xls..
        # wb_obj.protection.password = ''
        # After a fair amount of research, it seems that there is no easy
        # way to use a password withing Python to pass to an Excel file
        # May be easiest just to remove the password protection and re-enter
        # it when the work is done.

        # # print(wb_obj.sheetnames)
        this_sheet = wb_obj['Phone']
        # print(this_sheet)
        this_sheet.cell(row=ct, column=1).value = workerid
        this_sheet.cell(row=ct, column=2).value = phon_typ
        this_sheet.cell(row=ct, column=3).value = cntry
        this_sheet.cell(row=ct, column=4).value = intlcode
        this_sheet.cell(row=ct, column=5).value = area
        this_sheet.cell(row=ct, column=6).value = phon
        this_sheet.cell(row=ct, column=8).value = public

        wb_obj.save(csv_output + new_xl_file)

    except Exception as e:
        print("Error in phone_rec.py fn_ins_xl , Error = " + repr(e))
    # fn_write_error("Error in phone_rec.py - fn_get_id: "
    #                + repr(e))
    # fn_send_mail(settings.ADP_TO_EMAIL, settings.ADP_FROM_EMAIL,
    #          "Error in phone_rec.py fn_get_id, Error = " + repr(e),
    #          "Error in phone_rec.py fn_get_id")

def main():

    ##########################################################################
    # ==> python phone_rec.py --database=train --test
    # ==> python phone_rec.py --database=cars
    ##########################################################################

    # # Defines file names and directory location
    csv_output = settings.WORKDAY_CSV_OUTPUT
    csv_input = settings.WORKDAY_CSV_INPUT
    # print(csv_output)

    # For testing use last file
    # new_phone_file = csv_output + "ADPtoCXLast.csv"
    raw_phone_file = csv_input + "phones.csv"
    new_phone_file = csv_output + "phones_cln.csv"
    new_xl_file = "Worker Data.xlsx"

    fn_clear_sheet(csv_output, new_xl_file, 'Phone')

    try:
        # set global variable
        global EARL
        # determines which database is being called from the command line
        if database == 'cars':
            EARL = settings.INFORMIX_ODBC
        if database == 'train':
            EARL = settings.INFORMIX_ODBC_TRAIN
        else:
            # # this will raise an error when we call get_engine()
            # below but the argument parser should have taken
            # care of this scenario and we will never arrive here.
            EARL = None
            # establish database connection
        print(EARL)

        # Create the new csv file for the formatted data
        fn_write_phone_cl_header(new_phone_file)

        # Read the raw file and work on the formatting
        # print(raw_phone_file)
        with open(raw_phone_file, 'r') as f:
            d_reader = csv.DictReader(f, delimiter=',')
            ct = 2  # Leave at 1 to skip the header row...
            r = 0
            # Write to Excel spreadsheet

            for row in d_reader:
                # print("+++++++++++++++++++++")
                # print(row)
                r = r+1
                # print("ROW " + str(r))
                # print("Inserts " + str(ct))

                """Student workers and a few others typically 
                don't have the
                    Carthage ID Custom field populated  May have 
                    to force it
                    here"""
                if row['Carthage ID #'] == "":
                #     print('No Carth ID  ' + row['File Number'])
                    if row['File Number'] == "":
                        print("No id or file number")
                        pass
                    else:
                        # print(str(row['File Number']))
                        workerid = fn_get_id(str(row['File Number']), EARL,
                                             settings.INFORMIX_DEBUG)
                        # print(workerid)
                else:
                    workerid = row['Carthage ID #']

                """Format country and phone - I don't know if our 
                ADP  data
                    has any phone numbers that aren't in the US 
                    format"""
                cntry = fn_format_country(row['Primary Address: Country Code'])

                if len(row['Personal Contact: Home Phone']) != 0:
                    ret = fn_format_phone(row['Primary Address: Country Code'],
                                          row['Personal Contact: Home Phone'])
                    intlcode = ret[0]
                    area = ret[1]
                    phon = ret[2]
                    phon_typ = "Home Phone"
                    fn_write_clean_file(new_phone_file, [workerid + ','
                        + phon_typ + ',' + cntry + ',' + intlcode
                        + ','
                        + area + ',' + phon + ',' + '' + ',' + 'No'])
                    ct = ct + 1
                    fn_insert_xl(ct, workerid, phon_typ, cntry, intlcode, area,
                                 phon, 'No', csv_output, new_xl_file)

                if len(row['Personal Contact: Personal Mobile']) != 0:
                    ret = fn_format_phone(row['Primary Address: Country Code'],
                                   row['Personal Contact: Personal Mobile'])
                    intlcode = ret[0]
                    area = ret[1]
                    phon = ret[2]
                    phon_typ = "Personal Mobile"
                    fn_write_clean_file(new_phone_file, [workerid + ','
                        + phon_typ + ',' + cntry + ',' + intlcode
                        + ','
                        + area + ',' + phon + ',' + '' + ',' + 'No'])
                    ct = ct + 1
                    fn_insert_xl(ct, workerid, phon_typ, cntry, intlcode, area,
                                 phon, 'No', csv_output, new_xl_file)

                if len(row['Work Contact: Work Phone']) != 0:
                    ret = fn_format_phone(row['Primary Address: Country Code'],
                                          row['Work Contact: Work Phone'])
                    intlcode = ret[0]
                    area = ret[1]
                    phon = ret[2]
                    phon_typ = "Work Office"
                    fn_write_clean_file(new_phone_file, [workerid + ','
                        + phon_typ + ',' + cntry + ',' + intlcode
                        + ','
                        + area + ',' + phon + ',' + '' + ',' + 'No'])
                    ct = ct + 1
                    fn_insert_xl(ct, workerid, phon_typ, cntry, intlcode, area,
                                 phon, 'Yes', csv_output, new_xl_file)

                if len(row['Work Contact: Work Mobile']) != 0:
                    ret = fn_format_phone(
                        row['Primary Address: Country Code'],
                        row['Personal Contact: Work Mobile'])
                    intlcode = ret[0]
                    area = ret[1]
                    phon = ret[2]
                    phon_typ = "Work Mobile"
                    fn_write_clean_file(new_phone_file, [workerid + ','
                        + phon_typ + ',' + cntry + ',' + intlcode
                        + ','
                        + area + ',' + phon + ',' + '' + ',' + 'No'])
                    ct = ct + 1
                    fn_insert_xl(ct, workerid, phon_typ, cntry, intlcode,
                        area, phon, 'Yes', csv_output, new_xl_file)

        """Because we can't use the ADP info for student workers, we need to get 
                  that info from cx"""
        # Get the student data via SQL query
        ct = ct + 1

        stuquery = '''select distinct CR.adp_associate_id, CR.adp_id, JR.id, 
                trim(IR.phone), IR.ctry, JR.hrpay, JR.end_date, AR.aa, 
                trim(AR.phone)
                from job_rec JR
                LEFT JOIN id_rec IR
                ON JR.id = IR.id
                join cvid_rec CR
                on CR.cx_id = JR.id
                left join aa_rec AR 
                on AR.id = JR.id
                where (JR.end_date is null  or JR.end_date > TODAY)
                and JR.hrpay = 'DPW'
                and AR.aa in ('CELL', 'WORK')         
                limit 20
                  '''

        # print(stuquery)

        connection = get_connection(EARL)
        with connection:
            data_result = xsql(stuquery, connection).fetchall()
            ret = list(data_result)
            for i in ret:
                workerid = str(i[2])
                cntry = fn_format_cx_country(i[4])

                if len(i[3]) == 12:
                    ret = fn_format_cx_phone(i[4], i[3])
                    intlcode = ret[0]
                    phon = ret[2]
                    phon_typ = "Home Phone"
                    # print(workerid, phon_typ, cntry, intlcode, area, phon)
                    fn_write_clean_file(new_phone_file, [workerid + ','
                             + phon_typ + ',' + cntry + ',' + intlcode + ',' +
                             area + ',' + phon + ',' + '' + ',' + 'No'])

                    fn_insert_xl(ct, workerid, phon_typ, cntry, intlcode,
                             area, phon, 'No', csv_output, new_xl_file)
                    ct = ct + 1

                else:
                    # print('Phone invalid')
                    pass

                # print(str(len(i[8])))
                # print('Phone = ' + i[8])

                if len(i[8]) == 12:
                    ret = fn_format_cx_phone(i[4], i[8])
                    intlcode = ret[0]
                    phon = ret[2]
                    phon_typ = "Personal Mobile"
                    # print(workerid, phon_typ, cntry, intlcode, area, phon)
                    # print(ct)
                    fn_write_clean_file(new_phone_file, [workerid + ','
                             + phon_typ + ',' + cntry + ',' + intlcode + ',' +
                             area + ',' + phon + ',' + '' + ',' + 'No'])

                    fn_insert_xl(ct, workerid, phon_typ, cntry, intlcode,
                             area, phon, 'No', csv_output, new_xl_file)
                    ct = ct + 1
                else:
                    pass
                    # print('Phone invalid')

    except Exception as e:
        print("Error in phone_rec.py, Error = " + repr(e))
        # fn_write_error("Error in phone_rec.py - Main: "
        #                + repr(e))
        # fn_send_mail(settings.ADP_TO_EMAIL, settings.ADP_FROM_EMAIL,
        #          "Error in phone_rec.py, Error = " + repr(e),
        #          "Error in phone_rec.py")

if __name__ == "__main__":
    args = parser.parse_args()
    test = args.test
    database = args.database

    if not database:
        print("mandatory option missing: database name\n")
        exit(-1)
    else:
        database = database.lower()

    if database != 'cars' and database != 'train' and database != 'sandbox':
        print("database must be: 'cars' or 'train' or 'sandbox'\n")
        parser.print_help()

    sys.exit(main())
