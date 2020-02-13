import os
import sys
# import pysftp
import csv
import argparse
import logging
from logging.handlers import SMTPHandler
import django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "djmapache.settings.shell")
django.setup()
from django.conf import settings


import openpyxl
from openpyxl import Workbook
from django.conf import settings
from djimix.core.utils import get_connection, xsql
from djmapache.workday.hcm_hr.utilities import fn_format_country, \
    fn_write_personal_cl_header, fn_write_clean_file, fn_get_id


# informix environment
os.environ['INFORMIXSERVER'] = settings.INFORMIXSERVER
os.environ['DBSERVERNAME'] = settings.DBSERVERNAME
os.environ['INFORMIXDIR'] = settings.INFORMIXDIR
os.environ['ODBCINI'] = settings.ODBCINI
os.environ['ONCONFIG'] = settings.ONCONFIG
os.environ['INFORMIXSQLHOSTS'] = settings.INFORMIXSQLHOSTS
os.environ['LD_LIBRARY_PATH'] = settings.LD_LIBRARY_PATH
os.environ['LD_RUN_PATH'] = settings.LD_RUN_PATH

# # normally set as 'debug" in SETTINGS
# DEBUG = settings.INFORMIX_DEBUG

# set up command-line options
desc = """
    Upload ADP data to CX
"""
parser = argparse.ArgumentParser(description=desc)

parser.add_argument(
    "--test",
    action='store_true',
    help="Dry run?",
    dest="test"
)
parser.add_argument(
    "-d", "--database",
    help="database name.",
    dest="database"
)

def fn_clear_sheet(email_csv_output, new_xl_file):
    wb_obj = openpyxl.load_workbook(email_csv_output + new_xl_file)
    this_sheet = wb_obj['Personal']
    # print(this_sheet)
    row_count = this_sheet.max_row
    # col_count = this_sheet.max_column
    # print(row_count)
    # print(col_count)

    for row in this_sheet['A3:N' + str(row_count)]:
        # print(row)
        for cell in row:
            # print(cell.value)
            cell.value = ""
            # print(cell.value)

    wb_obj.save(email_csv_output + new_xl_file)

def fn_set_mrtl_stat(marital_code):
    switcher = {
        'S': 'Single',
        'M': 'Married',
        'D': 'Divorced',
        '': 'Not Disclosed',
        'P': 'Domestic Partner',
        'C': 'Civil Partnership',
        'E': 'Separated',
        'L': 'Legally Separated',
        'W': 'Widowed'
        }

    return switcher.get(marital_code, "Invalid marital status")


def fn_insert_xl(ct, workerid, bdate, ddate, gender, ctry,
                             mrtl_stat, mrtl_date, csv_output, new_xl_file):
    try:
        wb_obj = openpyxl.load_workbook(csv_output
                                        + new_xl_file)
        # this may be the syntax to open a pwd protected xls..
        # wb_obj.protection.password = ''
        # After a fair amount of research, it seems that there is no easy
        # way to use a password withing Python to pass to an Excel file
        # May be easiest just to remove the password protection and re-enter
        # it when the work is done.

        # # print(wb_obj.sheetnames)
        this_sheet = wb_obj['Personal']
        # print(this_sheet)
        this_sheet.cell(row=ct, column=1).value = workerid
        this_sheet.cell(row=ct, column=2).value = bdate
        this_sheet.cell(row=ct, column=3).value = ddate
        this_sheet.cell(row=ct, column=4).value = gender
        # Skip this column - calculated field
        # this_sheet.cell(row=ct, column=5).value = ctry
        this_sheet.cell(row=ct, column=6).value = mrtl_stat
        this_sheet.cell(row=ct, column=7).value = mrtl_date


        wb_obj.save(csv_output + new_xl_file)

    except Exception as e:
        print("Error in personal.py fn_ins_xl , Error = " + repr(e))
    # fn_write_error("Error in address.py - fn_get_id: "
    #                + repr(e))
    # fn_send_mail(settings.ADP_TO_EMAIL, settings.ADP_FROM_EMAIL,
    #          "Error in personal.py fn_get_id, Error = " + repr(e),
    #          "Error in personal.py fn_get_id")

def main():

    ##########################################################################
    # ==> python phone_rec.py --database=train --test
    # ==> python phone_rec.py --database=cars
    ##########################################################################

    # # Defines file names and directory location
    # if test:
    csv_output = "/home/dsullivan/djmapache/djmapache/workday/hcm_hr/" \
                            "clean_data/"
    csv_input = "/home/dsullivan/djmapache/djmapache/workday/hcm_hr/" \
                             "raw_data/"
    # else:
    # csv_output = settings.csv_output
    # print(csv_output)

    # For testing use last file
    # new_file = csv_output + "ADPtoCXLast.csv"
    raw_file = csv_input + "Personal.csv"
    new_file = csv_output + "personal_cln.csv"
    new_xl_file = "Worker Data.xlsx"

    fn_clear_sheet(csv_output, new_xl_file)

    try:
        # set global variable
        global EARL
        # determines which database is being called from the command line
        if database == 'cars':
            EARL = settings.INFORMIX_ODBC
        if database == 'train':
            EARL = settings.INFORMIX_ODBC_TRAIN
        else:
            # # this will raise an error when we call get_engine()
            # below but the argument parser should have taken
            # care of this scenario and we will never arrive here.
            EARL = None
            # establish database connection
        # print(EARL)

        # Pull the file from the ADP FTP site
        # execute sftp code in production only
        # if not test:
        #     file_download()

        # Create the new csv file for the formatted data
        fn_write_personal_cl_header(new_file)

        # Read the raw file and work on the formatting
        # print(raw_file)
        with open(raw_file, 'r') as f:
            d_reader = csv.DictReader(f, delimiter=',')
            ct = 2  # Leave at 1 to skip the header row...
            r = 0
            # Write to Excel spreadsheet

            for row in d_reader:
                # print("+++++++++++++++++++++")
                # print(row)
                r = r+1
                # print("ROW " + str(r))
                # print("Inserts " + str(ct))

                """Student workers and a few others typically
                don't have the
                    Carthage ID Custom field populated  May have
                    to force it
                    here"""
                if row['Carthage ID #'] == "":
                    # print('No Carth ID  ' + row['File Number'])
                    if row['File Number'] == "":
                        # print("No id or file number")
                        pass
                    else:
                        # print(str(row['File Number']))
                        workerid = fn_get_id(str(row['File Number']), EARL,
                                            settings.INFORMIX_DEBUG)
                        # print(workerid)
                else:
                    workerid = row['Carthage ID #']

                # print(workerid)

                # Associate ID
                # File Number
                # Carthage ID
                # # ","Birth Date","Date of Death","Gender Code","Gender",
                # "Primary Address: Country Code","Marital Status Code",
                # "Marital Status Effective Date","Position Status"

    #             # If ??? then employee else...
                bdate = row["Birth Date"]
                ddate = row["Date of Death"]
                gender = row["Gender"]
                gendrcod = row["Gender Code"]
                mrtl_stat = fn_set_mrtl_stat(row["Marital Status Code"].strip())
                mrtl_stat_cod = row["Marital Status Code"]
                mrtl_date = row["Marital Status Effective Date"]

                # print(new_file, [workerid + ',' + bdate
                #                                + ',' + ddate + ',' + gender
                #                                + ',' + mrtl_stat + ','
                #                                + mrtl_stat_cod + ','
                #                                + mrtl_date])

                fn_write_clean_file(new_file, [workerid + ',' + bdate
                    + ',' + ddate + ',' + gender + ',' + '' + ',' + mrtl_stat
                    + ',' + mrtl_date])

                ct = ct + 1

                fn_insert_xl(ct, workerid, bdate, ddate, gender, '',
                             mrtl_stat, mrtl_date,csv_output,  new_xl_file)

    #             else:
    #                 # print("No worker category")
    #                 # Probably need to write a log or notification here
    #                 pass
    #
    except Exception as e:
        print("Error in personal.py, Error = " + repr(e))
        # fn_write_error("Error in personal.py - Main: "
        #                + repr(e))
        # fn_send_mail(settings.ADP_TO_EMAIL, settings.ADP_FROM_EMAIL,
        #          "Error in personal.py, Error = " + repr(e),
        #          "Error in personal.py")

if __name__ == "__main__":
    args = parser.parse_args()
    test = args.test
    database = args.database

    if not database:
        print("mandatory option missing: database name\n")
        exit(-1)
    else:
        database = database.lower()

    if database != 'cars' and database != 'train' and database != 'sandbox':
        print("database must be: 'cars' or 'train' or 'sandbox'\n")
        parser.print_help()

    sys.exit(main())
